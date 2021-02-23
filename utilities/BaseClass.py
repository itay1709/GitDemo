import inspect
import logging

import openpyxl
import pytest
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

@pytest.mark.usefixtures("setupBrowser")
class BaseClass:
    def logScriptDemo(self):
        fileHandler = logging.FileHandler("logFile.log")  # define the file log name
        fileFormat = logging.Formatter("%(asctime)s :%(levelname)s :%(name)s :%(message)s")  # define the log format
        fileHandler.setFormatter(fileFormat)  # connect the format to the file

        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        logger.addHandler(fileHandler)  # method that needs to except as argument the file log name

        logger.setLevel(logging.DEBUG)

        return logger

    def waitUntilXpath(self, text):
        WebDriverWait(self.driver, 15).until(expected_conditions.presence_of_element_located((By.XPATH, text)))

    def selectByText(self, locator, text):
        sel = Select(locator)
        sel.select_by_visible_text(text)

    def getTestDataExcel(self, testCase):
        dic = {}
        testDataFile = openpyxl.load_workbook("C:\\Users\\P0022990\\Desktop\\Personal\\QA\Selenium python course\\testdata.xlsx")  # importfile
        sheet = testDataFile.active  # access the sheet
        for i in range(1, sheet.max_row + 1):  # iterate rows
            if sheet.cell(row=i, column=1).value == testCase:  # this for loop is use to extract specific test data for excel file
                for j in range(2, sheet.max_column + 1):  # iterate columns
                    dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                break
        return dic
