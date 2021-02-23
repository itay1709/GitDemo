import openpyxl
Dict = {}
testDataFile = openpyxl.load_workbook("C:\\Users\\P0022990\\Desktop\\Personal\\QA\Selenium python course\\testdata.xlsx")  #importfile
sheet = testDataFile.active  #access the sheet
firstName = sheet.cell(row=1, column=2)  #access to the relevant cell
print(firstName.value)
print(sheet['A2'].value)  #different way to access relevant cell
print(sheet.max_row)  #max row in sheet
print(sheet.max_column)  #max column in sheet

for i in range(1, sheet.max_row+1):  # iterate rows
    if sheet.cell(row=i, column=1).value == "test1":  # this for loop is use to extract specific test data for excel file
        for j in range(2, sheet.max_column+1):  # iterate columns
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        break
print(Dict)
